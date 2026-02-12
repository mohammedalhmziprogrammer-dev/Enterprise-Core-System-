import json
import pandas as pd
from io import BytesIO
from django.http import HttpResponse
from django.db.models import QuerySet, Model
from django.apps import apps
from typing import List, Dict, Any, Union, Optional
from datetime import datetime
import logging
import arabic_reshaper
from bidi.algorithm import get_display

logger = logging.getLogger(__name__)

class DynamicExporter:
    """
    محرك تصدير مركزي يعمل مع أي نموذج في النظام
    """
    
    @classmethod
    def export(
        cls,
        data_source: Union[str, QuerySet, List[Dict]],
        export_format: str = 'excel',
        columns: Optional[List[str]] = None,
        filters: Optional[Dict] = None,
        user=None,
        **kwargs
    ) -> HttpResponse:
        """
        تصدير البيانات لأي صيغة
        """
        try:
            # معالجة البيانات مباشرة بدون موديل
            if isinstance(data_source, list):
                return cls._export_raw_data(data_source, export_format, columns, **kwargs)
            
            # الحصول على البيانات
            data = cls._get_data(data_source, filters, user)
            
            # تحويل البيانات للشكل المطلوب
            formatted_data = cls._format_data(data, columns)
            
            # التصدير للصيغة المطلوبة
            if export_format == 'excel':
                return ExcelService.export(formatted_data, **kwargs)
            elif export_format == 'pdf':
                return ArabicPDFService.export(formatted_data, **kwargs)  # استخدم PDF عربي جديد
            elif export_format == 'csv':
                return CSVService.export(formatted_data, **kwargs)
            elif export_format == 'json':
                return JSONService.export(formatted_data, **kwargs)
            else:
                raise ValueError(f"صيغة {export_format} غير مدعومة")
                
        except Exception as e:
            logger.error(f"خطأ في التصدير: {str(e)}")
            raise
    
    @staticmethod
    def _export_raw_data(data: List[Dict], export_format: str, columns: Optional[List[str]] = None, **kwargs):
        """تصدير بيانات خام مباشرة"""
        if not data:
            raise ValueError("لا توجد بيانات للتصدير")
        
        # فلترة الأعمدة إذا تم تحديدها
        if columns and data:
            filtered_data = []
            for item in data:
                filtered_item = {k: v for k, v in item.items() if k in columns}
                filtered_data.append(filtered_item)
            data = filtered_data
        
        # التصدير للصيغة المطلوبة
        if export_format == 'excel':
            return ExcelService.export(data, **kwargs)
        elif export_format == 'pdf':
            return ArabicPDFService.export(data, **kwargs)
        elif export_format == 'csv':
            return CSVService.export(data, **kwargs)
        elif export_format == 'json':
            return JSONService.export(data, **kwargs)
        else:
            raise ValueError(f"صيغة {export_format} غير مدعومة")
    
    @staticmethod
    def _get_data(data_source, filters=None, user=None):
        """جلب البيانات من المصدر"""
        if isinstance(data_source, QuerySet):
            queryset = data_source
        elif isinstance(data_source, list):
            return data_source
        elif isinstance(data_source, str):
            # الحصول على الموديل من اسمه
            app_label, model_name = data_source.split('.')
            model = apps.get_model(app_label, model_name)
            queryset = model.objects.all()
        else:
            raise ValueError("مصدر البيانات غير صالح")
        
        # تطبيق الفلاتر إذا وجدت
        if filters:
            queryset = DynamicExporter._apply_filters(queryset, filters)
        
        # تطبيق صلاحيات المستخدم إذا كان موجوداً
        if user and hasattr(queryset.model, 'filter_by_user'):
            queryset = queryset.filter_by_user(user)
        
        return queryset
    
    @staticmethod
    def _apply_filters(queryset, filters):
        """تطبيق الفلاتر على QuerySet"""
        from django.db.models import Q
        
        q_objects = Q()
        
        for key, value in filters.items():
            if value is not None and value != '':
                if '__' in key:
                    q_objects &= Q(**{key: value})
                else:
                    q_objects &= Q(**{f"{key}__icontains": value})
        
        return queryset.filter(q_objects)
    
    @staticmethod
    def _format_data(data, columns=None):
        """
        تنسيق البيانات للتصدير
        Returns: قائمة من القواميس
        """
        if isinstance(data, QuerySet):
            # تحويل QuerySet إلى قواميس
            if columns:
                return list(data.values(*columns))
            return list(data.values())
        elif isinstance(data, list):
            if columns:
                # فلترة الأعمدة في القائمة
                return [{k: item.get(k) for k in columns if k in item} for item in data]
            return data
        
        raise ValueError("نوع البيانات غير مدعوم")
    
    @classmethod
    def get_model_fields(cls, model_path: str):
        """الحصول على حقول أي موديل"""
        try:
            app_label, model_name = model_path.split('.')
            model = apps.get_model(app_label, model_name)
            
            fields = []
            for field in model._meta.get_fields():
                if field.concrete and not field.many_to_many:
                    field_info = {
                        'name': field.name,
                        'verbose_name': getattr(field, 'verbose_name', field.name),
                        'type': field.get_internal_type(),
                        'choices': getattr(field, 'choices', None),
                    }
                    fields.append(field_info)
            
            return fields
        except Exception as e:
            logger.error(f"خطأ في الحصول على حقول الموديل: {str(e)}")
            return []



class ArabicPDFService:
    """خدمة PDF عربية 100% بدون تقطع"""
    
    @staticmethod
    def prepare_arabic_text(text):
        """تحضير النص العربي للعرض الصحيح"""
        if not isinstance(text, str):
            text = str(text)
        
        try:
            reshaped_text = arabic_reshaper.reshape(text)
            bidi_text = get_display(reshaped_text)
            return bidi_text
        except Exception as e:
            logger.warning(f"خطأ في معالجة النص العربي: {e}")
            return text
    
    @staticmethod
    def export(data, filename=None, title=None, **kwargs):
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, portrait
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            
            if not data or len(data) == 0:
                raise ValueError("لا توجد بيانات للتصدير")
            
            if not filename:
                filename = f"تصدير_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            elif not filename.endswith('.pdf'):
                filename = f"{filename}.pdf"
            
            buffer = BytesIO()
            
            try:
                pdfmetrics.registerFont(TTFont('Arial', 'Arial.ttf'))
                arabic_font = 'Arial'
            except:
                arabic_font = 'Helvetica'
            
            doc = SimpleDocTemplate(
                buffer,
                pagesize=portrait(A4),
                rightMargin=0.5*inch,
                leftMargin=0.5*inch,
                topMargin=0.5*inch,
                bottomMargin=0.5*inch,
                title=ArabicPDFService.prepare_arabic_text(title or 'تقرير')
            )
            
            elements = []
            styles = getSampleStyleSheet()
            
            arabic_title_style = ParagraphStyle(
                'ArabicTitle',
                parent=styles['Title'],
                fontName=arabic_font,
                fontSize=18,
                alignment=1,
                spaceAfter=12
            )
            
            arabic_normal_style = ParagraphStyle(
                'ArabicNormal',
                parent=styles['Normal'],
                fontName=arabic_font,
                fontSize=10,
                alignment=2,
                leading=12
            )
            
            if title:
                title_para = Paragraph(ArabicPDFService.prepare_arabic_text(title), arabic_title_style)
                elements.append(title_para)
                elements.append(Spacer(1, 12))
            
            date_para = Paragraph(
                ArabicPDFService.prepare_arabic_text(f"تاريخ التصدير: {datetime.now().strftime('%Y-%m-%d %H:%M')}"),
                arabic_normal_style
            )
            records_para = Paragraph(
                ArabicPDFService.prepare_arabic_text(f"عدد السجلات: {len(data)}"),
                arabic_normal_style
            )
            elements.append(date_para)
            elements.append(records_para)
            elements.append(Spacer(1, 20))
            
            # إعداد بيانات الجدول
            headers = list(data[0].keys())
            column_labels = kwargs.get('column_labels', {})
            display_headers = [ArabicPDFService.prepare_arabic_text(column_labels.get(h, h)) for h in headers]
            
            # تحويل رؤوس الجدول إلى Paragraph
            table_data = [[Paragraph(h, arabic_normal_style) for h in display_headers]]
            
            # تحضير البيانات مع Paragraph لكل خلية لضمان التفاف النصوص
            for row in data:
                row_data = []
                for header in headers:
                    value = row.get(header, '')
                    if isinstance(value, datetime):
                        value = value.strftime('%Y-%m-%d %H:%M')
                    elif value is None:
                        value = ''
                    prepared_value = ArabicPDFService.prepare_arabic_text(str(value))
                    row_data.append(Paragraph(prepared_value, arabic_normal_style))
                table_data.append(row_data)
            
            # حساب عرض الأعمدة ديناميكيًا حسب عدد الأعمدة
            page_width = A4[0] - (0.5 + 0.5) * inch  # عرض الصفحة ناقص الهوامش
            col_count = len(headers)
            col_widths = [page_width / col_count for _ in headers]  # توزيع متساوي
            
            table = Table(table_data, colWidths=col_widths, repeatRows=1)
            
            style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), arabic_font),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                ('TOPPADDING', (0, 0), (-1, 0), 10),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8f9fa')),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('FONTNAME', (0, 1), (-1, -1), arabic_font),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('ALIGN', (0, 1), (-1, -1), 'RIGHT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
                ('LEFTPADDING', (0, 0), (-1, -1), 6),
                ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                ('TOPPADDING', (0, 1), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
            ])
            
            table.setStyle(style)
            elements.append(table)
            
            # تذييل الصفحة
            elements.append(Spacer(1, 20))
            footer_para = Paragraph(
                ArabicPDFService.prepare_arabic_text(f"صفحة 1 - تم إنشاؤه بواسطة النظام - {datetime.now().strftime('%Y/%m/%d')}"),
                arabic_normal_style
            )
            elements.append(footer_para)
            
            doc.build(elements)
            buffer.seek(0)
            
            response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
            
        except ImportError as e:
            logger.error(f"مكتبة reportlab غير مثبتة: {e}")
            return ExcelService.export(data, filename, title, **kwargs)
        except Exception as e:
            logger.error(f"خطأ في إنشاء PDF: {str(e)}")
            return ExcelService.export(data, filename, title, **kwargs)


class ExcelService:
    """خدمة تصدير Excel"""
    
    @staticmethod
    def export(data, filename=None, title=None, **kwargs):
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        
        if not filename:
            filename = f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # إنشاء Workbook جديد
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # إضافة عنوان إذا وجد
        if title:
            ws.merge_cells('A1:D1')
            ws['A1'] = title
            ws['A1'].font = Font(size=14, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center')
            start_row = 3
        else:
            start_row = 1
        
        # كتابة الرؤوس
        if data and len(data) > 0:
            headers = list(data[0].keys())
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=start_row, column=col_num, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # كتابة البيانات
            for row_num, row_data in enumerate(data, start_row + 1):
                for col_num, header in enumerate(headers, 1):
                    value = row_data.get(header, '')
                    cell = ws.cell(row=row_num, column=col_num, value=value)
                    
                    # تنسيق التواريخ
                    if isinstance(value, datetime):
                        cell.number_format = 'YYYY-MM-DD HH:MM'
        
        # ضبط عرض الأعمدة
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # حفظ في Buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response


# للتوافق مع الكود القديم
class PDFService:
    """خدمة تصدير PDF (للتوافق)"""
    
    @staticmethod
    def export(data, filename=None, title=None, **kwargs):
        return ArabicPDFService.export(data, filename, title, **kwargs)


class AdvancedPDFService:
    """للتوافق مع الكود القديم"""
    
    @staticmethod
    def export(data, filename=None, title=None, **kwargs):
        return ArabicPDFService.export(data, filename, title, **kwargs)


class CSVService:
    """خدمة تصدير CSV"""
    
    @staticmethod
    def export(data, filename=None, **kwargs):
        import csv
        
        if not filename:
            filename = f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response.write(u'\ufeff'.encode('utf8'))  # BOM لـ UTF-8
        
        writer = csv.writer(response)
        
        if data and len(data) > 0:
            # كتابة الرؤوس
            headers = list(data[0].keys())
            writer.writerow(headers)
            
            # كتابة البيانات
            for row in data:
                writer.writerow([row.get(h, '') for h in headers])
        
        return response


class JSONService:
    """خدمة تصدير JSON"""
    
    @staticmethod
    def export(data, filename=None, **kwargs):
        if not filename:
            filename = f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        
        # تحويل QuerySet إلى قائمة
        if hasattr(data, 'values'):
            data = list(data.values())
        
        response = HttpResponse(
            json.dumps(data, ensure_ascii=False, indent=2, default=str),
            content_type='application/json'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response